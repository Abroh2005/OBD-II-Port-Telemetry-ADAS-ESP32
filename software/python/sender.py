# import sys, os, time, re, csv, glob, argparse
# from pathlib import Path

# try:
#     from openpyxl import load_workbook
# except Exception:
#     load_workbook = None

# NUM = re.compile(r"-?\d+(\.\d+)?")

# def to_float(s):
#     if s is None: return None
#     m = NUM.search(str(s))
#     return float(m.group(0)) if m else None

# def pick_idx(header, *cands):
#     h2i = {str(h).strip().lower(): i for i, h in enumerate(header)}
#     for c in cands:
#         if c.lower() in h2i:
#             return h2i[c.lower()]
#     return None

# def iter_xlsx_rows(path, sheet=None, max_rows=200000):
#     if load_workbook is None:
#         print("[ERR] openpyxl not installed. pip install openpyxl")
#         return
#     wb = load_workbook(filename=str(path), read_only=True, data_only=True)
#     ws = wb[sheet] if sheet and sheet in wb.sheetnames else wb.active
#     rows = ws.iter_rows(values_only=True)
#     header = next(rows, None)
#     if not header:
#         print("[ERR] Empty XLSX")
#         return

#     # Accept both naming styles:
#     # TIME, ENGINERPM, SPEED, THROTTLEPOS
#     # TIME, ENGINE_RPM, SPEED, THROTTLE_POS
#     iT = pick_idx(header, "TIME", "timestamp")
#     iR = pick_idx(header, "ENGINERPM", "ENGINE_RPM", "RPM", "EngineRPM")
#     iS = pick_idx(header, "SPEED", "Speed", "vehicle_speed")
#     iH = pick_idx(header, "THROTTLEPOS", "THROTTLE_POS", "ThrottlePos", "THROTTLE", "AccelPedal")

#     if None in (iT, iR, iS, iH):
#         print("[ERR] Missing required columns in XLSX header.")
#         print("Header:", header)
#         return

#     count = 0
#     for row in rows:
#         try:
#             t = int(to_float(row[iT]) or 0)
#             rpm = int(round(to_float(row[iR]) or 0))
#             spd = round(float(to_float(row[iS]) or 0.0), 1)
#             thr = round(float(to_float(row[iH]) or 0.0), 1)
#         except Exception:
#             continue
#         if 0 <= rpm <= 9000 and 0 <= spd <= 250 and 0 <= thr <= 100:
#             yield t, rpm, spd, thr
#             count += 1
#             if count >= max_rows:
#                 break

# def open_serial(port, baud):
#     import serial
#     return serial.Serial(port, baudrate=baud, timeout=0.1)

# def main():
#     ap = argparse.ArgumentParser(description="Stream dataset to ESP32 over Serial.")
#     ap.add_argument("--file", required=True)
#     ap.add_argument("--sheet", default=None)
#     ap.add_argument("--serial-port", required=True)
#     ap.add_argument("--baud", type=int, default=115200)
#     ap.add_argument("--rate-hz", type=float, default=10.0)
#     ap.add_argument("--loop", action="store_true")
#     args = ap.parse_args()

#     dataset = Path(args.file)
#     print(f"[INFO] Using dataset: {dataset}")

#     if dataset.suffix.lower() == ".xlsx":
#         row_iter_factory = lambda: iter_xlsx_rows(dataset, sheet=args.sheet)
#     else:
#         # CSV path retained from earlier version if needed
#         def iter_csv_rows(path):
#             with open(path, newline="", encoding="utf-8", errors="ignore") as f:
#                 reader = csv.reader(f)
#                 header = next(reader, None)
#                 if not header:
#                     print("[ERR] Empty CSV"); return
#                 iT = pick_idx(header, "TIME", "timestamp")
#                 iR = pick_idx(header, "ENGINERPM", "ENGINE_RPM", "RPM", "EngineRPM")
#                 iS = pick_idx(header, "SPEED", "Speed", "vehicle_speed")
#                 iH = pick_idx(header, "THROTTLEPOS", "THROTTLE_POS", "ThrottlePos", "THROTTLE", "AccelPedal")
#                 if None in (iT, iR, iS, iH):
#                     print("[ERR] Missing required columns in CSV header."); print("Header:", header); return
#                 for row in reader:
#                     try:
#                         t = int(to_float(row[iT]) or 0)
#                         rpm = int(round(to_float(row[iR]) or 0))
#                         spd = round(float(to_float(row[iS]) or 0.0), 1)
#                         thr = round(float(to_float(row[iH]) or 0.0), 1)
#                     except Exception:
#                         continue
#                     if 0 <= rpm <= 9000 and 0 <= spd <= 250 and 0 <= thr <= 100:
#                         yield t, rpm, spd, thr
#         row_iter_factory = lambda: iter_csv_rows(dataset)

#     # Prime rows
#     row_iter = row_iter_factory()
#     buffer = []
#     for _ in range(500):
#         try:
#             buffer.append(next(row_iter))
#         except StopIteration:
#             break
#         except Exception:
#             continue

#     if not buffer:
#         print("[ERR] No valid rows parsed from dataset.")
#         return
#     print(f"[INFO] Primed {len(buffer)} rows. Streaming...")

#     try:
#         ser = open_serial(args.serial_port, args.baud)
#         print(f"[INFO] Serial open on {args.serial_port} @ {args.baud}")
#     except Exception as e:
#         print(f"[ERR] Serial open failed: {e}")
#         return

#     period = 1.0 / max(0.1, args.rate_hz)
#     next_tick = time.time()
#     sent = 0
#     last_hb = time.time()

#     try:
#         while True:
#             if not buffer:
#                 row_iter = row_iter_factory()
#                 refilled = 0
#                 for _ in range(1000):
#                     try:
#                         buffer.append(next(row_iter))
#                         refilled += 1
#                     except StopIteration:
#                         break
#                     except Exception:
#                         continue
#                 if refilled == 0:
#                     if args.loop:
#                         continue
#                     print("[INFO] End of dataset."); break

#             t, rpm, spd, thr = buffer.pop(0)
#             line = f"t={t},rpm={rpm},speed={spd},thr={thr}"

#             now = time.time()
#             if now < next_tick:
#                 time.sleep(next_tick - now)
#             ser.write((line + "\n").encode("utf-8"))
#             print(line)
#             sent += 1
#             next_tick = time.time() + period

#             if time.time() - last_hb > 1.0:
#                 print(f"[HB] sent={sent}")
#                 last_hb = time.time()

#     except KeyboardInterrupt:
#         print("\n[INFO] Stopped.")
#     finally:
#         try: ser.close()
#         except: pass

# if __name__ == "__main__":
#     main()
import sys, os, time, re, csv, glob, argparse, socket
from pathlib import Path

try:
    from openpyxl import load_workbook
except Exception:
    load_workbook = None

NUM = re.compile(r"-?\d+(\.\d+)?")

def to_float(s):
    if s is None: return None
    m = NUM.search(str(s))
    return float(m.group(0)) if m else None

def pick_idx(header, *cands):
    h2i = {str(h).strip().lower(): i for i, h in enumerate(header)}
    for c in cands:
        if c.lower() in h2i:
            return h2i[c.lower()]
    return None

def iter_xlsx_rows(path, sheet=None, max_rows=200000):
    if load_workbook is None:
        print("[ERR] openpyxl not installed. pip install openpyxl")
        return
    wb = load_workbook(filename=str(path), read_only=True, data_only=True)
    ws = wb[sheet] if sheet and sheet in wb.sheetnames else wb.active
    rows = ws.iter_rows(values_only=True)
    header = next(rows, None)
    if not header:
        print("[ERR] Empty XLSX")
        return

    iT = pick_idx(header, "TIME", "timestamp")
    iR = pick_idx(header, "ENGINERPM", "ENGINE_RPM", "RPM", "EngineRPM")
    iS = pick_idx(header, "SPEED", "Speed", "vehicle_speed")
    iH = pick_idx(header, "THROTTLEPOS", "THROTTLE_POS", "ThrottlePos", "THROTTLE", "AccelPedal")

    if None in (iT, iR, iS, iH):
        print("[ERR] Missing required columns in XLSX header.")
        print("Header:", header)
        return

    count = 0
    for row in rows:
        try:
            t = int(to_float(row[iT]) or 0)
            rpm = int(round(to_float(row[iR]) or 0))
            spd = round(float(to_float(row[iS]) or 0.0), 1)
            thr = round(float(to_float(row[iH]) or 0.0), 1)
        except Exception:
            continue
        if 0 <= rpm <= 9000 and 0 <= spd <= 250 and 0 <= thr <= 100:
            yield t, rpm, spd, thr
            count += 1
            if count >= max_rows:
                break

def iter_csv_rows(path):
    with open(path, newline="", encoding="utf-8", errors="ignore") as f:
        reader = csv.reader(f)
        header = next(reader, None)
        if not header:
            print("[ERR] Empty CSV"); return
        iT = pick_idx(header, "TIME", "timestamp")
        iR = pick_idx(header, "ENGINERPM", "ENGINE_RPM", "RPM", "EngineRPM")
        iS = pick_idx(header, "SPEED", "Speed", "vehicle_speed")
        iH = pick_idx(header, "THROTTLEPOS", "THROTTLE_POS", "ThrottlePos", "THROTTLE", "AccelPedal")
        if None in (iT, iR, iS, iH):
            print("[ERR] Missing required columns in CSV header.")
            print("Header:", header); return
        for row in reader:
            try:
                t = int(to_float(row[iT]) or 0)
                rpm = int(round(to_float(row[iR]) or 0))
                spd = round(float(to_float(row[iS]) or 0.0), 1)
                thr = round(float(to_float(row[iH]) or 0.0), 1)
            except Exception:
                continue
            if 0 <= rpm <= 9000 and 0 <= spd <= 250 and 0 <= thr <= 100:
                yield t, rpm, spd, thr

def open_serial(port, baud):
    import serial
    return serial.Serial(port, baudrate=baud, timeout=0.1)

def main():
    ap = argparse.ArgumentParser(
        description="Stream dataset to ESP32 via CAN (Serial→Nano) or WiFi UDP.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
EXAMPLES:
  WiFi Only:
    python sender.py --file data.xlsx --sheet origicleaned --mode wifi --wifi-ip 10.99.148.174 --rate-hz 2 --loop

  Serial Only (via Nano):
    python sender.py --file data.xlsx --sheet origicleaned --mode serial --serial-port COM12 --rate-hz 2 --loop

  Both (Dual transmission):
    python sender.py --file data.xlsx --sheet origicleaned --mode both --serial-port COM12 --wifi-ip 10.99.148.174 --rate-hz 2 --loop
        """)

    ap.add_argument("--file", required=True, help="Dataset file (XLSX or CSV)")
    ap.add_argument("--sheet", default=None, help="XLSX sheet name (default: first sheet)")
    ap.add_argument("--mode", choices=["serial", "wifi", "both"], default="wifi", 
                    help="Transmission mode (default: wifi)")
    ap.add_argument("--serial-port", default=None, help="Serial port (e.g., COM12). Required for serial/both modes")
    ap.add_argument("--baud", type=int, default=115200, help="Serial baud rate (default: 115200)")
    ap.add_argument("--wifi-ip", default="10.99.148.174", help="ESP32 WiFi IP address (default: 10.99.148.174)")
    ap.add_argument("--wifi-port", type=int, default=4210, help="ESP32 UDP port (default: 4210)")
    ap.add_argument("--rate-hz", type=float, default=2.0, help="Send rate in Hz (default: 2.0)")
    ap.add_argument("--loop", action="store_true", help="Loop dataset when finished")

    args = ap.parse_args()

    # Validate arguments
    if args.mode in ["serial", "both"] and args.serial_port is None:
        print("[ERR] --serial-port required for serial/both modes")
        return

    dataset = Path(args.file)
    if not dataset.exists():
        print(f"[ERR] Dataset not found: {dataset}")
        return

    print(f"[INFO] Mode: {args.mode.upper()}")
    print(f"[INFO] Dataset: {dataset}")

    if dataset.suffix.lower() == ".xlsx":
        row_iter_factory = lambda: iter_xlsx_rows(dataset, sheet=args.sheet)
    else:
        row_iter_factory = lambda: iter_csv_rows(dataset)

    # Prime rows
    row_iter = row_iter_factory()
    buffer = []
    for _ in range(500):
        try:
            buffer.append(next(row_iter))
        except StopIteration:
            break
        except Exception:
            continue

    if not buffer:
        print("[ERR] No valid rows in dataset")
        return

    print(f"[INFO] Primed {len(buffer)} rows. Streaming...\n")

    # Setup serial
    ser = None
    if args.mode in ["serial", "both"]:
        try:
            ser = open_serial(args.serial_port, args.baud)
            print(f"[Serial] Open on {args.serial_port} @ {args.baud} baud")
        except Exception as e:
            print(f"[ERR] Serial failed: {e}")
            if args.mode == "serial":
                return
            print("[WARN] Continuing without serial\n")

    # Setup WiFi UDP
    udp_sock = None
    if args.mode in ["wifi", "both"]:
        try:
            udp_sock = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
            print(f"[WiFi] Sending to {args.wifi_ip}:{args.wifi_port}")
        except Exception as e:
            print(f"[ERR] WiFi UDP failed: {e}")
            if args.mode == "wifi":
                return
            print("[WARN] Continuing without WiFi\n")

    if not ser and not udp_sock:
        print("[ERR] No valid transmission method available")
        return

    print(f"[INFO] Rate: {args.rate_hz} Hz ({1/args.rate_hz:.1f}s period)\n")
    print("-" * 80)

    period = 1.0 / max(0.1, args.rate_hz)
    next_tick = time.time()
    sent = 0
    last_hb = time.time()

    try:
        while True:
            # Refill buffer
            if not buffer:
                row_iter = row_iter_factory()
                refilled = 0
                for _ in range(1000):
                    try:
                        buffer.append(next(row_iter))
                        refilled += 1
                    except StopIteration:
                        break
                    except Exception:
                        continue

                if refilled == 0:
                    if args.loop:
                        print("\n[INFO] End of dataset - looping...")
                        continue
                    else:
                        print("\n[INFO] End of dataset - stopping")
                        break

            # Get and send data
            t, rpm, spd, thr = buffer.pop(0)
            line = f"t={t},rpm={rpm},speed={spd},thr={thr}"

            # Rate limit
            now = time.time()
            if now < next_tick:
                time.sleep(next_tick - now)

            # Send via serial (to Nano)
            if ser:
                try:
                    ser.write((line + "\n").encode("utf-8"))
                except:
                    pass

            # Send via WiFi UDP (to ESP32)
            if udp_sock:
                try:
                    udp_sock.sendto((line + "\n").encode("utf-8"), (args.wifi_ip, args.wifi_port))
                except:
                    pass

            print(line)
            sent += 1
            next_tick = time.time() + period

            # Heartbeat
            if time.time() - last_hb > 2.0:
                print(f"[HB] Sent: {sent} packets | Mode: {args.mode.upper()}")
                last_hb = time.time()

    except KeyboardInterrupt:
        print("\n[INFO] Stopped by user")
    finally:
        try:
            if ser: ser.close()
            print("[Serial] Closed")
        except:
            pass
        try:
            if udp_sock: udp_sock.close()
            print("[WiFi] Closed")
        except:
            pass
        print(f"\n[INFO] Total sent: {sent} packets")

if __name__ == "__main__":
    main()
